from __future__ import annotations
import logging
import os

import openpyxl

logging.basicConfig(filename="TimesheetParser.log", level=logging.DEBUG, format='%(asctime)s :: %(levelname)s :: %(message)s')

class TimesheetParser:
    """
    class to parse data from list of .xlsx files (of particular format) that contain timesheets and form .xlsx report
    """

    def __init__(self: TimesheetParser, filenames: list[str]) -> None:
        """
        :param filenames: list of file names of .xlsx files that contain timesheets
        :raises FileNotFoundError: if there is no file named filename
        """
        self.workbook_list = list()
        for filename in filenames:
            self.workbook_list.append(openpyxl.load_workbook(filename=filename)) # automatically raises FileNotFoundError if there is no such file

    def get_report_by_project_name(self: TimesheetParser, project_name: str, project_pseudonyms: list[str]=[]) -> dict[dict[str, float]]:
        """
        :param project_name: string that represents project name in timesheet
        :param project_pseudonyms: list of pseudonyms that can be used as an alternative project name (mainly for legacy support)
        :returns report: report is a dictionary in format {section1: {employee1: hours, ...}, section2: {employee1: hours, ...}, ...}
        """
        report = dict()
        for workbook in self.workbook_list:
            for sheet in workbook:
                try:
                    row_start, column_start, row_end, column_end = self.find_timesheet_table_coordinates_by_sheet(sheet)
                    project_number_column_number = 2 # currently column named "Project number" has number 2
                    section_of_project_column_number = 3 # currently column named "Section of project" has number 3
                    days_column_number = 1 # currently column named "Working days (%)" has number 1
                    number_of_rows_in_table = row_end - row_start + 1
                    for i in range(number_of_rows_in_table):
                        if (sheet.cell(row=(row_start + i), column=(column_start + project_number_column_number)).value in [project_name] + project_pseudonyms):    # if this cell is under "Project number" column
                            try: # tryng to add time if there is already existing record of the employee for the section
                                print(f"section: {sheet.cell(row=(row_start + i), column=(column_start + section_of_project_column_number)).value}, employee: {sheet.title}, report: {report}")
                                if (sheet.cell(row=(row_start + i), column=(column_start + section_of_project_column_number)).value == None):
                                    report["section not given"][sheet.title] += sheet.cell(row=(row_start + i), column=(column_start + days_column_number)).value
                                else:
                                    report[sheet.cell(row=(row_start + i), column=(column_start + section_of_project_column_number)).value][sheet.title] += sheet.cell(row=(row_start + i), column=(column_start + days_column_number)).value
                            except KeyError as ex:
                                logging.info(f"During adding working days, {ex} happened.")
                                if (sheet.cell(row=(row_start + i), column=(column_start + section_of_project_column_number)).value == None):
                                    if (not "section not given" in report.keys()):
                                        report["section not given"] = dict()
                                    report["section not given"][sheet.title] = sheet.cell(row=(row_start + i), column=(column_start + days_column_number)).value
                                else:
                                    if (not sheet.cell(row=(row_start + i), column=(column_start + section_of_project_column_number)).value in report.keys()):
                                        report[sheet.cell(row=(row_start + i), column=(column_start + section_of_project_column_number)).value] = dict()
                                    report[sheet.cell(row=(row_start + i), column=(column_start + section_of_project_column_number)).value][sheet.title] = sheet.cell(row=(row_start + i), column=(column_start + days_column_number)).value
                except ValueError as ex:    # catching exceptions for cases when function was not able to determine timesheet table coordinates
                    logging.info(f"During generating report, {ex} happened.")
        return report

    @staticmethod
    def find_timesheet_table_coordinates_by_sheet(sheet: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, int, int, int]:
        """
        different timesheets can have different coordinates of start and end, this function searches for table coordinates based on top-right, top-left, bottom-right and bottom-left cell values.
        :param sheet: openpyxl sheet that has timesheet table
        :raises ValueError: if table coordinates cannot be found
        :returns coordinates: tuple of four coordinates of timesheet table: row_start, row_end, column_start, column_end
        """
        # setting variables to None in order to be able to check if they had been set after search 
        row_start = None
        row_end = None
        column_start = None
        column_end = None

        # timesheet table borders are defined by top-right, top-left, bottom-right and bottom-left cell values
        # right now, top-left has "Day" and cell on the right has "Working days(%)" / "Hours worked(%)" / "Number of hours"
        # top-right has "Scope of design work(short description of performed activities)" and cell on the right has "Type of work"
        # bottom-left has "Approval by the supervising person"
        for row in sheet.iter_rows():
            for cell in row:
                if (type(cell.value) != str):
                    continue
                hours_possible_column_names = ["workingdays(%)", "numberofhours", "hoursworked(%)"]
                if ( (cell.value.replace(" ", "").lower() == "day") and (sheet.cell(row=cell.row, column=cell.column + 1).value.replace(" ", "").lower() in hours_possible_column_names) ):
                    row_start = cell.row
                    column_start = cell.column
                if ( (cell.value.replace(" ", "").lower() == '''Scope of design work
                    (short description of performed activities)'''.replace(" ", "").lower()) and (sheet.cell(row=cell.row, column=cell.column + 1).value.replace(" ", "").lower() == "Тype of work".replace(" ", "").lower()) ):
                    column_end = cell.column
                if (cell.value.replace(" ", "").lower() == "Approval by the supervising person".replace(" ", "").lower()):
                    row_end = cell.row
        
        if (row_start == None or column_start == None):  # if start coordinates of table hadn't been found
            raise ValueError(f"Start of timesheet table hadn't been found. Sheet: {sheet.title}")
        
        if (row_end == None or column_end == None):
            raise ValueError(f"End of timesheet table hadn't been found. Sheet: {sheet.title}")

        return (row_start, column_start, row_end, column_end)

    @staticmethod
    def write_report_to_xlsx_file(report: dict[dict[str, float]], filename: str) -> None:
        """
        method to write report dict to .xlsx file named filename
        :param report: report is a dictionary in format {section1: {employee1: hours, ...}, section2: {employee1: hours, ...}, ...}
        """
        basepath = os.path.dirname(__file__)                            # get absolute path to this file
        filepath = os.path.abspath(os.path.join(basepath, "..", "resources", "output", f"{filename}.xlsx"))
        workbook = openpyxl.Workbook()
        workbook.remove(workbook["Sheet"])  # remove standard workbook sheet "Sheet"
        sheet_name = "Fact"
        workbook.create_sheet(sheet_name)

        table_begin_row = 4
        table_begin_column = 2

        counter = 0     # to count rows
        for section in report.keys():
            workbook[sheet_name].cell(row=(table_begin_row + counter), column=table_begin_column).value = section
            for employee in report[section]:
                workbook[sheet_name].cell(row=(table_begin_row + counter), column=(table_begin_column + 1)).value = employee
                workbook[sheet_name].cell(row=(table_begin_row + counter), column=(table_begin_column + 2)).value = report[section][employee]
                counter += 1

        workbook.save(filepath)

if __name__ == "__main__":
    obj = TimesheetParser(["test_3_month.xlsx"])
    report = obj.get_report_by_project_name("Issyk-Kul 1. ДУ1")
    print(report)
    obj.write_report_to_xlsx_file(report, "test")
#    except Exception as ex:
 #       print(f"{ex} happened")
